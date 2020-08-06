# 参考: https://openpyxl.readthedocs.io/en/stable/tutorial.html#create-a-workbook
import datetime
from openpyxl import Workbook

def main(workbook):
    # 当前的表格
    currentWorkSheet = workbook.active
    # 当前表格名字
    currentWorkSheet.title = "测试xlsx"
    currentWorkSheet['A1'] = 'test'
    currentWorkSheet.sheet_properties.tabColor = "1072BA"
    # 创建一个有名称表格 Workbook.create_sheet(name?: string, 需要插入的 number)
    workssheet1 = workbook.create_sheet("测试数据")

    currentWorkSheet['A1'] = datetime.datetime(2010, 7, 21)
    currentWorkSheet['A1'].number_format

    for row in range(1, 40):
        workssheet1.append(range(10))


def init(): 
    # 创建一个文件  一个文件有多个表格
    workbook = Workbook()
    # 操作表格
    main(workbook)
    # 保存文件
    workbook.save('./files/workbook2.xlsx')

# init
init()






# When a worksheet is created in memory, it contains no cells. They are created when first accessed.
# 创建表格时不会创建任何一格， 只有访问他们的时候才创建。

# 直接通过名称获取某个表格
# ws3 = wb["New Title"]
# 打印出所有的表格名称
# >>> print(wb.sheetnames)
# ['Sheet2', 'New Title', 'Sheet1']
# 复制表格
# >>> source = wb.active
# >>> target = wb.copy_worksheet(source)
# 访问某一格
# >>> d = ws.cell(row=4, column=2, value=10)


### 访问区间
# >>> colC = ws['C']
# >>> col_range = ws['C:D']
# >>> row10 = ws[10]
# >>> row_range = ws[5:10]


### 遍历数据
# You can also use the Worksheet.iter_rows() method:

# >>> for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
# ...    for cell in row:
# ...        print(cell)


# Likewise the Worksheet.iter_cols() method will return columns:

# >>> for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
# ...     for cell in col:
# ...         print(cell)

### 访问所有行和列
# tuple(ws.rows)
# tuple(ws.columns)


### 访问每一个的值
# for row in ws.values:
#        for value in row:
#      print(value)


### 读取一个已经存在的表格
# wb2 = load_workbook('test.xlsx')
# print wb2.sheetnames


### 访问具体一格(字母:列, 数字:行)

# >>> c = ws['A4']
# >>> ws['A4'] = 4
# >>> d = ws.cell(row=4, column=2, value=10)



# >>> # set date using a Python datetime
# >>> ws['A1'] = datetime.datetime(2010, 7, 21)
# >>>
# >>> ws['A1'].number_format



# 时间
# currentWorkSheet['A1'] = datetime.datetime(2010, 7, 21)
# currentWorkSheet['A1'].number_format

# 公式
# >>> # add a simple formula
# >>> ws["A1"] = "=SUM(1, 1)"





# 只读模式
# wb = load_workbook(filename='large_file.xlsx', read_only=True)
# >>> wb = Workbook(write_only=True)



# 增删改查

# >>> ws.insert_rows(7)
# >>> ws.delete_cols(6, 3)
# >>> ws.move_range("D4:F10", rows=-1, cols=2)





